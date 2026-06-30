#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx_to_prolog.py
=================
把 antimalarial_drug_enrichment.py 产出的 xlsx 转成 SWI-Prolog 知识库 (.pl)。

设计原则（针对 Plasmodium 知识图谱 / CDSS）：
  · 药名 -> 规范 Prolog atom（小写、下划线、去括注）：Cipargamin (KAE609) -> cipargamin
  · 多值字段拆成多条事实，便于查询/推理：
        lifecycle_stages "liver; hypnozoite; gametocyte"
            -> acts_on_stage(primaquine, liver).
               acts_on_stage(primaquine, hypnozoite).
               acts_on_stage(primaquine, gametocyte).
        indications 同理拆 drug_indication/2
  · 分类 / 机理 / 靶标带【来源】参数，保留可溯源性：
        drug_class(chloroquine, review, '4-Aminoquinoline').
        drug_class(chloroquine, atc, 'Aminoquinolines (P01BA)').
        drug_moa(chloroquine, chembl, 'Heme polymerase inhibitor').
        drug_target(sj733, 'PfATP4（Na⁺-ATPase）', 'review(MoA)').
  · 数值字段不加引号（weight/alogp/psa/...），Prolog 可直接算术比较。
  · 空字段不输出事实（含当前为空的 resistance_mechanism 占位）。
  · 文件头 :- encoding(utf8). 保证中文 atom 正常 consult。

用法：
    python xlsx_to_prolog.py                       # 默认读 antimalarial_drugs_enriched.xlsx
    python xlsx_to_prolog.py 输入.xlsx 输出.pl
依赖：
    pip install openpyxl
在 SWI-Prolog 中加载：
    ?- consult('antimalarial_drugs.pl').
    ?- acts_on_stage(D, hypnozoite).          % 哪些药杀休眠体（防复发）
    ?- drug_class(D, atc, C).                  % 按 WHO ATC 分类
"""

import re
import sys
from datetime import datetime

# --------------------------------------------------------------------------- #
# 列 -> 谓词 的映射配置
# --------------------------------------------------------------------------- #
# 标量元数据：drug_attr(Drug, Key, Value)（值加引号）
ATTR_FIELDS = {
    "phase": "phase", "trade_name": "trade_name", "fda_year": "approval_year",
    "modality": "modality", "status_proxy": "dev_status", "action_type": "action_type",
    "iupac_name": "iupac_name",
}
# 分类：drug_class(Drug, Source, Value)
CLASS_FIELDS = {
    "review_class": "review", "cn_class": "cn", "atc_class_who": "atc",
    "cf_parent": "classyfire_parent", "cf_subclass": "classyfire_subclass",
    "cf_class": "classyfire_class", "cf_superclass": "classyfire_superclass",
}
# 机理：drug_moa(Drug, Source, Text)
MOA_FIELDS = {"review_moa": "review", "chembl_moa": "chembl"}
# ADMET：drug_admet(Drug, Key, Number)
ADMET_FIELDS = {"alogp": "alogp", "psa": "psa", "hbd": "hbd",
                "hba": "hba", "ro5_violations": "ro5_violations"}
# 化学结构：drug_chem(Drug, Key, Value)；NUMERIC_CHEM 内的值不加引号
CHEM_FIELDS = {
    "chemical_formula": "formula", "weight_average": "weight_average",
    "monoisotopic_mass": "monoisotopic_mass", "exact_mass": "exact_mass",
    "smiles_isomeric": "smiles_isomeric", "smiles_canonical": "smiles_canonical",
    "inchi": "inchi", "inchikey": "inchikey",
}
NUMERIC_CHEM = {"weight_average", "monoisotopic_mass", "exact_mass"}
# 数据库交叉引用：drug_xref(Drug, DB, Id)
XREF_FIELDS = {"drugbank_id": "drugbank", "chembl_id": "chembl", "pubchem_cid": "pubchem"}

# 声明为 discontiguous 的谓词（按药物分组输出，故需声明）
PREDICATES = [
    "drug/1", "drug_label/2", "drug_status/2", "drug_attr/3",
    "drug_class/3", "drug_moa/3", "drug_action_type/2",
    "drug_target/3", "drug_target_xref/3", "drug_target_organism/2",
    "acts_on_stage/2", "drug_stage_raw/2",
    "drug_resistance/3",
    "drug_warning/2", "drug_indication/2", "drug_admet/3",
    "drug_chem/3", "drug_xref/3", "drug_note/2",
]


# --------------------------------------------------------------------------- #
# 工具
# --------------------------------------------------------------------------- #
def atomize(name: str) -> str:
    """药名 -> 合法 Prolog atom（无需引号的形式）。"""
    name = re.sub(r"\s*\(.*?\)\s*", " ", str(name)).strip()   # 去括注
    name = name.lower()
    name = re.sub(r"[^a-z0-9]+", "_", name).strip("_")
    if not name:
        return "unknown"
    if name[0].isdigit():
        name = "d_" + name
    return name


def patom(s) -> str:
    """任意值 -> 单引号 Prolog atom（转义反斜杠与单引号、压平换行）。"""
    s = str(s).replace("\\", "\\\\").replace("'", "\\'")
    s = s.replace("\n", " ").replace("\r", " ").strip()
    return "'" + s + "'"


def as_number(s):
    """能转成数字就返回 int/float，否则 None。"""
    if isinstance(s, bool):
        return None
    if isinstance(s, (int, float)):
        return s
    t = str(s).strip()
    if re.fullmatch(r"-?\d+", t):
        return int(t)
    if re.fullmatch(r"-?\d*\.\d+", t):
        return float(t)
    return None


def nonempty(v) -> bool:
    return v is not None and str(v).strip() not in ("", "无", "None")


# 生命周期阶段归一化：自由文本 -> 标准 atom
_STAGE_RULES = [
    (("blood", "无性", "asexual", "玫瑰花结", "隔离"), "blood_asexual"),
    (("liver", "肝", "因果预防"),                      "liver"),
    (("hypnozoite", "休眠"),                           "hypnozoite"),
    (("gametocyte", "配子"),                           "gametocyte"),
    (("sporozoite", "子孢子"),                         "sporozoite"),
    (("辅助", "宿主导向", "host"),                      "host_directed"),
]


def normalize_stages(raw: str) -> list[str]:
    """把 lifecycle_stages 文本拆分并归一为标准阶段 atom 列表（去重保序）。"""
    out = []
    for token in re.split(r"[;；,，/]", str(raw)):
        tok = token.strip()
        if not tok:
            continue
        matched = None
        for keys, atom in _STAGE_RULES:
            if any(k in tok for k in keys):
                matched = atom
                break
        if matched is None:
            matched = atomize(tok)          # 未识别则降级为 sanitized atom
        if matched not in out:
            out.append(matched)
    return out


# --------------------------------------------------------------------------- #
# 单药 -> 事实
# --------------------------------------------------------------------------- #
def emit_drug(row: dict) -> list[str]:
    name = row.get("display_name", "")
    d = atomize(name)
    L = [f"\n% ===== {name} ({d}) ====="]
    L.append(f"drug({d}).")
    L.append(f"drug_label({d}, {patom(name)}).")

    status = (row.get("status") or "").strip().lower()
    if status:
        L.append(f"drug_status({d}, {status if status in ('approved','candidate') else patom(status)}).")

    # 标量元数据
    for col, key in ATTR_FIELDS.items():
        v = row.get(col)
        if nonempty(v) and str(v).strip() != "—":
            num = as_number(v) if key == "approval_year" else None
            val = num if num is not None else patom(v)
            L.append(f"drug_attr({d}, {key}, {val}).")

    # 分类（带来源）
    for col, src in CLASS_FIELDS.items():
        v = row.get(col)
        if nonempty(v) and not str(v).startswith("无"):
            L.append(f"drug_class({d}, {src}, {patom(v)}).")

    # 机理（带来源）
    for col, src in MOA_FIELDS.items():
        v = row.get(col)
        if nonempty(v):
            L.append(f"drug_moa({d}, {src}, {patom(v)}).")

    # 靶标（molecular_target + 来源）
    tgt = row.get("molecular_target")
    if nonempty(tgt):
        src = row.get("stage_target_src") or "unspecified"
        L.append(f"drug_target({d}, {patom(tgt)}, {patom(src)}).")
    # ChEMBL 策展靶标作为另一来源
    ctgt = row.get("target_name")
    if nonempty(ctgt):
        L.append(f"drug_target({d}, {patom(ctgt)}, chembl).")
    if nonempty(row.get("target_organism")):
        L.append(f"drug_target_organism({d}, {patom(row['target_organism'])}).")
    for col, db in (("target_uniprot", "uniprot"), ("target_chembl_id", "chembl")):
        v = row.get(col)
        if nonempty(v):
            for piece in re.split(r"[;,]", str(v)):
                piece = piece.strip()
                if piece:
                    L.append(f"drug_target_xref({d}, {db}, {patom(piece)}).")
    if nonempty(row.get("action_type")):
        L.append(f"drug_action_type({d}, {patom(row['action_type'])}).")

    # 生命周期阶段（多值拆分 + 原文保真）
    stages_raw = row.get("lifecycle_stages")
    if nonempty(stages_raw):
        L.append(f"drug_stage_raw({d}, {patom(stages_raw)}).")
        for st in normalize_stages(stages_raw):
            L.append(f"acts_on_stage({d}, {st}).")

    # 耐药机制（当前多为空占位；非空才输出，带来源）
    res = row.get("resistance_mechanism")
    if nonempty(res):
        rsrc = row.get("resistance_source") or "unspecified"
        L.append(f"drug_resistance({d}, {patom(res)}, {patom(rsrc)}).")

    # 警告
    if nonempty(row.get("drug_warning")):
        for w in re.split(r"\s*\|\s*", str(row["drug_warning"])):
            if w.strip():
                L.append(f"drug_warning({d}, {patom(w.strip())}).")
    # 适应症（多值拆分）
    if nonempty(row.get("indications")):
        for ind in re.split(r"[;；]", str(row["indications"])):
            ind = ind.strip()
            if ind:
                L.append(f"drug_indication({d}, {patom(ind)}).")

    # ADMET（数值）
    for col, key in ADMET_FIELDS.items():
        v = row.get(col)
        num = as_number(v)
        if num is not None:
            L.append(f"drug_admet({d}, {key}, {num}).")

    # 化学结构
    for col, key in CHEM_FIELDS.items():
        v = row.get(col)
        if not nonempty(v):
            continue
        if key in NUMERIC_CHEM:
            num = as_number(v)
            L.append(f"drug_chem({d}, {key}, {num if num is not None else patom(v)}).")
        else:
            L.append(f"drug_chem({d}, {key}, {patom(v)}).")

    # 数据库交叉引用
    for col, db in XREF_FIELDS.items():
        v = row.get(col)
        if nonempty(v):
            num = as_number(v)
            L.append(f"drug_xref({d}, {db}, {num if (db=='pubchem' and num is not None) else patom(v)}).")

    # 备注
    if nonempty(row.get("notes")):
        L.append(f"drug_note({d}, {patom(row['notes'])}).")
    return L


# --------------------------------------------------------------------------- #
# 读取 xlsx
# --------------------------------------------------------------------------- #
def read_xlsx(path: str) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("缺少 openpyxl：pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(rows)]
    out = []
    for r in rows:
        if r is None or all(c is None for c in r):
            continue
        out.append({header[i]: r[i] for i in range(len(header))})
    return out


# --------------------------------------------------------------------------- #
def main() -> int:
    in_path = sys.argv[1] if len(sys.argv) > 1 else "antimalarial_drugs_enriched_v6_25_EN.xlsx"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "plasmodium_antimalarial_drugs.pl"

    rows = read_xlsx(in_path)
    lines = [
        ":- encoding(utf8).",
        f"% 抗疟药物知识库 —— 由 xlsx_to_prolog.py 自动生成",
        f"% 源文件: {in_path}",
        f"% 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"% 药物条目: {len(rows)}",
        "% 请勿手工编辑；改数据请改 xlsx 后重跑转换器。",
        "",
    ]
    lines += [f":- discontiguous {p}." for p in PREDICATES]

    fact_count = 0
    for row in rows:
        facts = emit_drug(row)
        fact_count += sum(1 for f in facts if not f.lstrip().startswith("%") and f.strip())
        lines += facts

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    print(f"✓ 已生成 {out_path}")
    print(f"  药物 {len(rows)} 个，事实约 {fact_count} 条")
    print(f"  在 SWI-Prolog 中：?- consult('{out_path}').")
    return 0


if __name__ == "__main__":
    sys.exit(main())